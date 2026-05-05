"""
Generate realistic FinOps sample files for testing the chat upload feature.

Run from repo root with the project venv active:
    python demo-data/generate.py

Produces (in demo-data/):
  azure-cost-90days.csv         ~30k rows, Cost Management-shaped
  azure-resources.tsv           ~5k rows, Resource Graph inventory
  advisor-recommendations.json  ~500 items, JSON array
  cost-export.json              nested object with summary + breakdown
  finops-notes.md               long markdown narrative
  audit.log                     ~20k log lines
  vm-inventory.xlsx             3 sheets (VMs, Disks, Snapshots)
  finops-report.pdf             multi-page text PDF
  cost-export.parquet           ~50k rows
"""
from __future__ import annotations
import csv
import json
import os
import random
from datetime import date, timedelta
from pathlib import Path

random.seed(42)
OUT = Path(__file__).parent

SUBSCRIPTIONS = [
    ("11111111-1111-1111-1111-111111111111", "prod-platform"),
    ("22222222-2222-2222-2222-222222222222", "prod-data"),
    ("33333333-3333-3333-3333-333333333333", "dev-sandbox"),
    ("44444444-4444-4444-4444-444444444444", "shared-services"),
]
RGS = ["rg-platform-weu", "rg-data-eus2", "rg-aks-prod", "rg-network-hub", "rg-mgmt", "rg-ml-train", "rg-sql-prod"]
SERVICES = [
    ("Virtual Machines", 0.34),
    ("Storage", 0.18),
    ("Azure SQL", 0.12),
    ("Azure Kubernetes Service", 0.10),
    ("App Service", 0.07),
    ("Cosmos DB", 0.05),
    ("Bandwidth", 0.04),
    ("Log Analytics", 0.04),
    ("Azure Monitor", 0.02),
    ("Cognitive Services", 0.02),
    ("Container Registry", 0.01),
    ("Key Vault", 0.01),
]
LOCATIONS = ["westeurope", "northeurope", "eastus", "eastus2", "swedencentral", "uksouth"]
TAGS_ENV = ["prod", "dev", "test", "stage", None]
TAGS_CC = ["eng", "data", "platform", "ml", None]


def weighted_choice(pairs):
    total = sum(w for _, w in pairs)
    r = random.random() * total
    upto = 0
    for v, w in pairs:
        upto += w
        if upto >= r:
            return v
    return pairs[-1][0]


def gen_csv():
    p = OUT / "azure-cost-90days.csv"
    start = date.today() - timedelta(days=90)
    rows = 0
    with p.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "UsageDate", "SubscriptionId", "SubscriptionName", "ResourceGroup",
            "ServiceName", "MeterCategory", "ResourceLocation", "ResourceId",
            "Quantity", "UnitPrice", "PreTaxCost", "Currency", "Tags",
        ])
        for d in range(90):
            day = start + timedelta(days=d)
            for _ in range(350):
                sub_id, sub_name = random.choice(SUBSCRIPTIONS)
                rg = random.choice(RGS)
                svc = weighted_choice(SERVICES)
                loc = random.choice(LOCATIONS)
                qty = round(random.uniform(0.1, 24.0), 4)
                unit = round(random.uniform(0.001, 0.85), 6)
                # weekly weekend dip + small upward drift
                drift = 1 + (d / 90) * 0.18
                weekend = 0.78 if day.weekday() >= 5 else 1.0
                cost = round(qty * unit * drift * weekend * random.uniform(0.85, 1.2), 4)
                resource_name = f"{svc.lower().replace(' ', '-')[:20]}-{random.randint(1000, 9999)}"
                rid = f"/subscriptions/{sub_id}/resourceGroups/{rg}/providers/Microsoft.{svc.split()[0]}/{resource_name}"
                env = random.choice(TAGS_ENV)
                cc = random.choice(TAGS_CC)
                tags_parts = []
                if env: tags_parts.append(f"environment={env}")
                if cc: tags_parts.append(f"cost-center={cc}")
                tags = ";".join(tags_parts)
                w.writerow([day.isoformat(), sub_id, sub_name, rg, svc, svc, loc, rid, qty, unit, cost, "USD", tags])
                rows += 1
    print(f"  CSV    {p.name:32} rows={rows:>6}  size={p.stat().st_size/1024/1024:.2f} MB")


def gen_tsv():
    p = OUT / "azure-resources.tsv"
    rows = 5000
    with p.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter="\t")
        w.writerow(["id", "name", "type", "resourceGroup", "subscriptionId", "location", "skuName", "skuTier", "powerState", "tags"])
        types = [
            ("Microsoft.Compute/virtualMachines", ["Standard_B2s", "Standard_D4s_v5", "Standard_E8s_v5", "Standard_FX8ms_v2"]),
            ("Microsoft.Compute/disks", ["P10", "P20", "P30", "S20"]),
            ("Microsoft.Storage/storageAccounts", ["Standard_LRS", "Standard_ZRS", "Premium_LRS"]),
            ("Microsoft.Sql/servers/databases", ["GP_S_Gen5_2", "GP_Gen5_4", "BC_Gen5_8"]),
            ("Microsoft.ContainerService/managedClusters", ["Standard_D4s_v5", "Standard_D8s_v5"]),
            ("Microsoft.Network/publicIPAddresses", ["Basic", "Standard"]),
            ("Microsoft.Network/networkInterfaces", ["-"]),
            ("Microsoft.Web/sites", ["P0v3", "P1v3", "P2v3"]),
        ]
        for i in range(rows):
            t, skus = random.choice(types)
            sku = random.choice(skus)
            sub_id, _ = random.choice(SUBSCRIPTIONS)
            rg = random.choice(RGS)
            name = f"{t.split('/')[-1][:6]}-{i:05d}"
            ps = random.choice(["running", "running", "running", "stopped", "deallocated"]) if "virtualMachines" in t else "-"
            env = random.choice(TAGS_ENV) or ""
            cc = random.choice(TAGS_CC) or ""
            tags = f"environment={env};cost-center={cc}".strip(";")
            w.writerow([
                f"/subscriptions/{sub_id}/resourceGroups/{rg}/providers/{t}/{name}",
                name, t, rg, sub_id, random.choice(LOCATIONS), sku,
                "Standard" if "Standard" in sku else "Basic", ps, tags,
            ])
    print(f"  TSV    {p.name:32} rows={rows:>6}  size={p.stat().st_size/1024:.1f} KB")


def gen_advisor_json():
    p = OUT / "advisor-recommendations.json"
    items = []
    categories = ["Cost", "Cost", "Cost", "Performance", "Reliability", "Security", "OperationalExcellence"]
    impacts = ["High", "Medium", "Low"]
    short = [
        ("Right-size or shutdown underutilized virtual machines", "Standard_FX8ms_v2 < 5% CPU over 7 days"),
        ("Buy reserved virtual machine instances to save over your pay-as-you-go costs", "3-year RI on Standard_D4s_v5"),
        ("Delete unattached disks", "Premium SSD P30 disk unattached for 47 days"),
        ("Use Standard Storage to store Managed Disks snapshots", "Snapshots on premium storage"),
        ("Configure Auto-Scale for App Service Plan", "Plan at <20% CPU 24h"),
        ("Enable Azure Hybrid Benefit for Windows VMs", "12 eligible VMs without AHUB"),
        ("Reduce Log Analytics ingestion via table-level retention", "Heartbeat 90d retention"),
        ("Migrate to Standard SSD from Premium SSD where IOPS allow", "Disk has <500 IOPS p95"),
    ]
    for i in range(500):
        sub_id, _ = random.choice(SUBSCRIPTIONS)
        rg = random.choice(RGS)
        cat = random.choice(categories)
        imp = random.choice(impacts)
        short_d, ext = random.choice(short)
        savings = round(random.uniform(8, 980), 2) if cat == "Cost" else 0
        items.append({
            "id": f"/subscriptions/{sub_id}/resourceGroups/{rg}/providers/Microsoft.Advisor/recommendations/{i:05x}",
            "category": cat,
            "impact": imp,
            "shortDescription": short_d,
            "extendedProperties": {"reason": ext, "annualSavingsAmount": savings, "savingsCurrency": "USD"},
            "resourceMetadata": {
                "resourceId": f"/subscriptions/{sub_id}/resourceGroups/{rg}/providers/Microsoft.Compute/virtualMachines/vm-{i:05d}",
                "subscriptionId": sub_id,
                "resourceGroup": rg,
                "location": random.choice(LOCATIONS),
            },
            "lastUpdated": (date.today() - timedelta(days=random.randint(0, 30))).isoformat(),
        })
    p.write_text(json.dumps(items, indent=2))
    print(f"  JSON   {p.name:32} items={len(items):>6}  size={p.stat().st_size/1024:.1f} KB")


def gen_cost_export_json():
    p = OUT / "cost-export.json"
    obj = {
        "schemaVersion": "1.0",
        "billingPeriod": "2026-04",
        "currency": "USD",
        "totals": {"actualCost": 184_532.41, "amortizedCost": 168_220.07, "forecastNextMonth": 191_005.10},
        "byService": [
            {"serviceName": s[0], "cost": round(184532.41 * s[1] * random.uniform(0.9, 1.1), 2)} for s in SERVICES
        ],
        "bySubscription": [
            {"subscriptionId": sid, "name": sname, "cost": round(random.uniform(15000, 80000), 2)}
            for sid, sname in SUBSCRIPTIONS
        ],
        "anomalies": [
            {"date": (date.today() - timedelta(days=i)).isoformat(),
             "service": random.choice([s[0] for s in SERVICES]),
             "delta": round(random.uniform(50, 1500), 2)}
            for i in range(1, 12)
        ],
        "tagBreakdown": {
            "cost-center": {"eng": 78230.10, "data": 41200.55, "platform": 32008.40, "ml": 15920.20, "untagged": 17173.16},
            "environment": {"prod": 142010.20, "dev": 21005.05, "test": 8312.16, "untagged": 13205.00},
        },
    }
    p.write_text(json.dumps(obj, indent=2))
    print(f"  JSON   {p.name:32} keys={len(obj):>6}  size={p.stat().st_size/1024:.1f} KB")


def gen_md():
    p = OUT / "finops-notes.md"
    sections = [
        "# Q2 FinOps Review — Notes\n\n",
        "## Goals\n- Cut waste by 12% MoM\n- Move 60% of stable VMs onto 1y RIs\n- Tag coverage > 95% for cost-center & environment\n\n",
        "## Findings\n",
    ]
    for i in range(120):
        sections.append(
            f"### Finding {i+1}: {random.choice(['Idle VM cluster', 'Unattached disks', 'Oversized SQL DB', 'Untagged storage', 'Missing budget', 'Premium disk on test workload'])}\n"
            f"- Resource group: `{random.choice(RGS)}`\n"
            f"- Estimated monthly waste: **${random.randint(40, 980)}**\n"
            f"- Owner: {random.choice(['Cloud CoE', 'Platform team', 'Data team', 'ML team'])}\n"
            f"- Action: {random.choice(['Delete after 7d notice', 'Right-size next maintenance window', 'Apply autoshutdown 19:00 UTC', 'Move to Standard SSD', 'Add cost-center tag'])}\n\n"
        )
    sections.append("## Next Steps\n1. Review with finance week of May 10\n2. Publish chargeback dashboard\n3. Roll out Azure Policy for required tags\n")
    p.write_text("".join(sections))
    print(f"  MD     {p.name:32} chars={p.stat().st_size:>6}  size={p.stat().st_size/1024:.1f} KB")


def gen_log():
    p = OUT / "audit.log"
    levels = ["INFO", "INFO", "INFO", "WARN", "ERROR"]
    actors = ["alice@contoso.com", "bob@contoso.com", "ci-pipeline", "system"]
    actions = [
        "tag.applied environment=prod",
        "budget.created amount=5000 scope=subscription",
        "vm.deallocated reason=autoshutdown",
        "rbac.assigned role=Reader",
        "policy.evaluated compliant=false",
        "anomaly.detected service=Storage delta=320.50",
        "export.scheduled storageAccount=stcostsprodweu",
        "savingsplan.recommendation generated savings=4120.00",
    ]
    with p.open("w", encoding="utf-8") as f:
        for i in range(20000):
            ts = (date.today() - timedelta(days=random.randint(0, 30))).isoformat()
            f.write(f"{ts}T{random.randint(0,23):02d}:{random.randint(0,59):02d}:{random.randint(0,59):02d}Z "
                    f"{random.choice(levels):5} corrId=req-{random.randint(10000,99999):05d} "
                    f"actor={random.choice(actors):24} {random.choice(actions)}\n")
    print(f"  LOG    {p.name:32} lines={20000:>6}  size={p.stat().st_size/1024:.1f} KB")


def gen_xlsx():
    try:
        from openpyxl import Workbook
    except ImportError:
        print("  XLSX   skipped (openpyxl not installed)")
        return
    p = OUT / "vm-inventory.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "VMs"
    ws.append(["Name", "ResourceGroup", "Subscription", "Size", "OS", "PowerState", "MonthlyCost", "CpuP95", "MemP95"])
    for i in range(800):
        sub_id, sub_name = random.choice(SUBSCRIPTIONS)
        ws.append([
            f"vm-{i:04d}", random.choice(RGS), sub_name,
            random.choice(["Standard_B2s", "Standard_D4s_v5", "Standard_E8s_v5", "Standard_FX8ms_v2"]),
            random.choice(["Linux", "Windows"]),
            random.choice(["running", "running", "stopped", "deallocated"]),
            round(random.uniform(20, 1200), 2),
            round(random.uniform(2, 95), 1),
            round(random.uniform(10, 88), 1),
        ])
    ws2 = wb.create_sheet("Disks")
    ws2.append(["Name", "ResourceGroup", "Sku", "SizeGB", "AttachedTo", "MonthlyCost"])
    for i in range(600):
        ws2.append([
            f"disk-{i:04d}", random.choice(RGS),
            random.choice(["P10", "P20", "P30", "S20"]),
            random.choice([128, 256, 512, 1024, 2048]),
            f"vm-{random.randint(0, 800):04d}" if random.random() > 0.18 else "",
            round(random.uniform(5, 180), 2),
        ])
    ws3 = wb.create_sheet("Snapshots")
    ws3.append(["Name", "ResourceGroup", "SizeGB", "AgeDays", "MonthlyCost"])
    for i in range(300):
        ws3.append([f"snap-{i:04d}", random.choice(RGS), random.choice([128, 256, 512]),
                    random.randint(1, 400), round(random.uniform(2, 48), 2)])
    wb.save(p)
    print(f"  XLSX   {p.name:32} sheets=3       size={p.stat().st_size/1024:.1f} KB")


def gen_pdf():
    try:
        from reportlab.lib.pagesizes import LETTER
        from reportlab.pdfgen import canvas
    except ImportError:
        # Fallback to a tiny hand-written PDF (no deps) so testing PDF still works.
        p = OUT / "finops-report.pdf"
        body = ("BT /F1 12 Tf 50 750 Td (Azure FinOps Report - Q2 2026) Tj ET\n"
                "BT /F1 10 Tf 50 730 Td (Total spend: $184,532) Tj ET\n"
                "BT /F1 10 Tf 50 715 Td (Top service: Virtual Machines 34%) Tj ET\n"
                "BT /F1 10 Tf 50 700 Td (Recommended savings: $42,108/month) Tj ET\n")
        # Minimal PDF skeleton
        pdf = (
            b"%PDF-1.4\n"
            b"1 0 obj <</Type /Catalog /Pages 2 0 R>> endobj\n"
            b"2 0 obj <</Type /Pages /Kids [3 0 R] /Count 1>> endobj\n"
            b"3 0 obj <</Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Contents 4 0 R /Resources <</Font <</F1 5 0 R>>>>>> endobj\n"
            b"4 0 obj <</Length " + str(len(body)).encode() + b">> stream\n" + body.encode() + b"endstream endobj\n"
            b"5 0 obj <</Type /Font /Subtype /Type1 /BaseFont /Helvetica>> endobj\n"
            b"xref\n0 6\n0000000000 65535 f \n0000000010 00000 n \n0000000060 00000 n \n0000000110 00000 n \n0000000220 00000 n \n0000000400 00000 n \ntrailer <</Size 6 /Root 1 0 R>>\nstartxref\n470\n%%EOF\n"
        )
        p.write_bytes(pdf)
        print(f"  PDF    {p.name:32} (minimal, install reportlab for multi-page)")
        return
    p = OUT / "finops-report.pdf"
    c = canvas.Canvas(str(p), pagesize=LETTER)
    width, height = LETTER
    for page in range(8):
        c.setFont("Helvetica-Bold", 16)
        c.drawString(72, height - 72, f"Azure FinOps Report — Q2 2026 (page {page+1}/8)")
        c.setFont("Helvetica", 10)
        y = height - 110
        lines = [
            "Executive Summary",
            "",
            f"Total spend last 30 days: ${random.randint(150_000, 220_000):,}",
            f"Forecast next 30 days:  ${random.randint(160_000, 235_000):,}",
            "",
            "Top cost drivers",
            "  1. Virtual Machines  34%  ($62,741)",
            "  2. Storage           18%  ($33,216)",
            "  3. Azure SQL         12%  ($22,144)",
            "  4. AKS               10%  ($18,453)",
            "",
            "Recommendations",
        ]
        for i in range(20):
            lines.append(f"  - {random.choice(['Right-size','Reserved','Delete idle','Apply tag','Set budget'])} action #{page*20+i+1}: save ~${random.randint(10, 950)}/mo")
        for line in lines:
            c.drawString(72, y, line)
            y -= 14
            if y < 72: break
        c.showPage()
    c.save()
    print(f"  PDF    {p.name:32} pages=8        size={p.stat().st_size/1024:.1f} KB")


def gen_parquet():
    try:
        import pandas as pd
    except ImportError:
        print("  PARQUET skipped (pandas not installed)")
        return
    p = OUT / "cost-export.parquet"
    n = 50_000
    df = pd.DataFrame({
        "UsageDate": [date.today() - timedelta(days=random.randint(0, 90)) for _ in range(n)],
        "SubscriptionId": [random.choice(SUBSCRIPTIONS)[0] for _ in range(n)],
        "ResourceGroup": [random.choice(RGS) for _ in range(n)],
        "ServiceName": [weighted_choice(SERVICES) for _ in range(n)],
        "Location": [random.choice(LOCATIONS) for _ in range(n)],
        "Quantity": [round(random.uniform(0.1, 24), 4) for _ in range(n)],
        "UnitPrice": [round(random.uniform(0.001, 0.85), 6) for _ in range(n)],
        "PreTaxCost": [round(random.uniform(0.01, 240), 4) for _ in range(n)],
        "Currency": ["USD"] * n,
    })
    try:
        df.to_parquet(p, index=False)
    except Exception as e:
        print(f"  PARQUET skipped ({e})")
        return
    print(f"  PARQUET {p.name:31} rows={n:>6}  size={p.stat().st_size/1024:.1f} KB")


if __name__ == "__main__":
    print(f"Generating sample files in {OUT}")
    gen_csv()
    gen_tsv()
    gen_advisor_json()
    gen_cost_export_json()
    gen_md()
    gen_log()
    gen_xlsx()
    gen_pdf()
    gen_parquet()
    print("Done.")
